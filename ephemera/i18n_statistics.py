"""
Стата по str_i18n, html_i18n
"""
import sys
import os
import openpyxl as px

sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

from visual import create_app
from visual.models import Footage, Tour


app = create_app('config.local.py')

with app.app_context():
    tour_meta_keys = [
        'actions'
        'active_meshes',
        'blotch',
        'navigator',
        'overlays',
        'splash',
        'toolbar',
        'skyboxes',
    ]
    footage_meta_keys = [
        'floors',
        'skyboxes',
    ]
    counter = {
        'Action.title': {'count':0, 'tours': []},
        'Action.body': {'count':0, 'tours': []},
        'Action.html': {'count':0, 'tours': []},
        'Action.buttons.text': {'count':0, 'tours': []},
        'TourSkybox.title': {'count':0, 'tours': []},
        'Overlay.tooltip': {'count':0, 'tours': []},
        'OverlayTextWidget.text': {'count':0, 'tours': []},
        'Blotch.text': {'count':0, 'tours': []},
        'Blotch.html': {'count':0, 'tours': []},
        'Splash.cancel_text': {'count':0, 'tours': []},
        'Splash.cancel_html': {'count':0, 'tours': []},
        'ActiveMesh.title': {'count':0, 'tours': []},
        'NavigatorElement.title': {'count':0, 'tours': []},
        'ToolbarButton.title': {'count':0, 'tours': []},
        'ToolbarButton.text': {'count':0, 'tours': []},
        'Floor.title': {'count':0, 'tours': []},
        'FootageSkybox.title': {'count':0, 'tours': []}
    }

    def add_stat(name, tour_id):
        if tour_id not in counter[name]['tours']:
            counter[name]['tours'].append(tour_id)
            counter[name]['count'] += 1

    tours = Tour.query.\
        join(Footage, Tour.footage_id == Footage.id).\
        filter(Footage.type.in_(['virtual', 'real'])).\
        all()

    for tour in tours:
        tour_meta = tour.meta
        footage_meta = tour.footage.meta

        for meta_key in tour_meta_keys:
            if meta_key == 'actions':
                for key, action in tour_meta.get(meta_key, {}).items():
                    if 'title' in action and type(action['title']) is dict:
                        add_stat('Action.title', tour.id)
                    if 'body' in action and type(action['body']) is dict:
                        add_stat('Action.body', tour.id)
                    if 'htnl' in action and type(action['html']) is dict:
                        add_stat('Action.html', tour.id)
                    if 'buttons' in action:
                        for button in action['buttons']:
                            if 'text' in button and type(button['text']) is dict:
                                add_stat('Action.buttons.text', tour.id)
            if meta_key == 'skyboxes':
                for key, skybox in tour_meta.get(meta_key, {}).items():
                    if 'title' in skybox and type(skybox['title']) is dict:
                        add_stat('TourSkybox.title', tour.id)
            if meta_key == 'overlays':
                for key, overlay in tour_meta.get(meta_key, {}).items():
                    if 'tooltip' in overlay and type(overlay['tooltip']) is dict:
                        add_stat('Overlay.tooltip', tour.id)
                    if 'widget' in overlay and 'text' in overlay['widget']:
                        if type(overlay['widget']['text']) is dict:
                            add_stat('OverlayTextWidget.text', tour.id)
            if meta_key == 'blotch':
                if meta_key in tour_meta and tour_meta[meta_key] is not None and 'text' in tour_meta[meta_key]:
                    if type(tour_meta[meta_key]['text']) is dict:
                        add_stat('Blotch.text', tour.id)
                if meta_key in tour_meta and tour_meta[meta_key] is not None and 'html' in tour_meta[meta_key]:
                    if type(tour_meta[meta_key]['html']) is dict:
                        add_stat('Blotch.html', tour.id)
            if meta_key == 'splash':
                if meta_key in tour_meta and 'cancel_text' in tour_meta[meta_key]:
                    if type(tour_meta[meta_key]['cancel_text']) is dict:
                        add_stat('Splash.cancel_text', tour.id)
                if meta_key in tour_meta and 'cancel_html' in tour_meta[meta_key]:
                    if type(tour_meta[meta_key]['cancel_html']) is dict:
                        add_stat('Splash.cancel_html', tour.id)
            if meta_key == 'active_meshes':
                for key, active_mesh in tour_meta.get(meta_key, {}).items():
                    if 'title' in active_mesh and type(active_mesh['title']) is dict:
                        add_stat('ActiveMesh.title', tour.id)
            if meta_key == 'navigator':
                for navigator_element in tour_meta.get(meta_key, []):
                    if 'title' in navigator_element and type(navigator_element['title']) is dict:
                        add_stat('NavigatorElement.title', tour.id)
            if meta_key == 'toolbar':
                for toolbar_element in tour_meta.get(meta_key, []):
                    if 'title' in toolbar_element and type(toolbar_element['title']) is dict:
                        add_stat('ToolbarButton.title', tour.id)
                    if 'text' in toolbar_element and type(toolbar_element['text']) is dict:
                        add_stat('ToolbarButton.text', tour.id)

        for meta_key in footage_meta_keys:
            if meta_key == 'floors':
                for key, floor in footage_meta.get(meta_key, {}).items():
                    if 'title' in floor and type(floor['title']) is dict:
                        add_stat('Floor.title', tour.id)
            if meta_key == 'skyboxes':
                for key, skybox in footage_meta.get(meta_key, {}).items():
                    if 'title' in skybox and type(skybox['title']) is dict:
                        add_stat('FootageSkybox.title', tour.id)


wb = px.Workbook()
dest_filename = 'i18n_stat.xlsx'
ws1 = wb.active

count = 1

ws1.cell(row=count, column=1, value='Свойство')
ws1.cell(row=count, column=2, value='Количество туров, в котором оно интернационализировано')
ws1.cell(row=count, column=3, value='%')
ws1.cell(row=count, column=4, value='Туры')

for prop, stat in counter.items():
        count += 1
        ws1.cell(row=count, column=1, value=prop)
        ws1.cell(row=count, column=2, value=stat['count'])
        ws1.cell(row=count, column=3, value='%3.5f' % float(100*stat['count']/len(tours)))
        ws1.cell(row=count, column=4, value=','.join(str(x) for x in stat['tours']))

wb.save(filename=dest_filename)



